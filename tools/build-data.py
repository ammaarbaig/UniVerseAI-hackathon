#!/usr/bin/env python3
"""
tools/build-data.py

Reads 2026.08-College-Database-final.xlsx and writes assets/js/data.js as
window.UNIVERSE_DATA. Deterministic: no random, no timestamps, no
dict/set-iteration-order dependence (every collection is explicitly sorted
before it is written out).

Re-run any time the workbook changes:
    python tools/build-data.py
"""

import json
import os
import re
import zlib

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK_PATH = os.path.join(ROOT, "2026.08-College-Database-final.xlsx")
OUT_PATH = os.path.join(ROOT, "assets", "js", "data.js")

VERIFIED_ON = "12 Aug 2026"

CATEGORIES = ["General", "EWS", "OBC-NCL", "SC", "ST"]
BRANCHES = ["CSE", "IT", "ECE", "Electrical", "Mechanical", "Civil", "Chemical", "AI & DS"]

BRANCH_FACTOR = {
    "CSE": 1.00,
    "IT": 1.35,
    "AI & DS": 1.15,
    "ECE": 1.70,
    "Electrical": 2.20,
    "Mechanical": 2.60,
    "Civil": 3.10,
    "Chemical": 3.40,
}

CATEGORY_MULTIPLIER = {
    "General": 1.00,
    "EWS": 1.17,
    "OBC-NCL": 1.33,
    "SC": 2.36,
    "ST": 2.98,
}

TYPE_BASE = {
    "State University": 6000,
    "Government": 9000,
    "Private": 22000,
}

INSTITUTION_TYPE_MAP = {
    "Government / DTE": "Government",
    "Private": "Private",
    "Government/state-funded": "State University",
}

# Free-text college labels used in 07_Cutoffs that do not match
# 02_Institutions_Master.Institution_Name verbatim. Resolved by hand against
# 04_Engineering_REAP_Ref (city + name are unambiguous for these).
CUTOFF_COLLEGE_ALIAS = {
    "MBM University, Jodhpur": "REAP-001",
    "GEC Bikaner": "REAP-006",
    "GEC Ajmer": "REAP-005",
    "GEC Barmer": "REAP-012",
    "GEC Bharatpur": "REAP-010",
    "GEC Jhalawar": "REAP-007",
    "GEC Banswara": "REAP-011",
    "MLVTEC Bhilwara": "REAP-008",
    "JECRC": "REAP-016",
    "Arya College of Engineering & IT": "REAP-019",
    "SKIT Jaipur": "REAP-015",
    "Poornima College of Engineering": "REAP-017",
    "College of Technology and Engineering, Udaipur": "REAP-002",
    "Government Engineering College, Ajmer": "REAP-005",
    # "GEC Baran" has no counterpart in 04_Engineering_REAP_Ref — deliberately
    # left unmapped, its cutoff rows are dropped.
}

CUTOFF_BRANCH_ALIAS = {
    "CSE": "CSE",
    "IT": "IT",
    "Information Technology": "IT",
    "ECE": "ECE",
    "Electrical": "Electrical",
    "Mechanical": "Mechanical",
    "Civil": "Civil",
    "Chemical": "Chemical",
    "AI & DS": "AI & DS",
    # "Electronics & Computer Engineering" and "Mechanical Engineering" (the
    # GECA row, whose value is unparseable anyway) are intentionally absent.
}


# ───────────────────────── text cleanup ─────────────────────────

_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_WHITESPACE = re.compile(r"[ \t]+")


def clean(value):
    """Replace mojibake (U+FFFD) and stray control chars; collapse whitespace."""
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    s = value.replace("�", "-")
    s = _CONTROL_CHARS.sub("", s)
    s = _WHITESPACE.sub(" ", s)
    s = s.strip()
    return s if s != "" else None


def norm(s):
    """Lowercase, alnum-only, collapsed-whitespace form used for name matching."""
    if s is None:
        return ""
    s = str(s).lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def norm_status(s):
    if not s:
        return "SECONDARY"
    u = s.upper()
    if u.startswith("PRIMARY"):
        return "PRIMARY"
    return "SECONDARY"


# ───────────────────────── workbook access ─────────────────────────


def sheet_rows(wb, name):
    """Yield each data row (below the header) as a dict of cleaned values,
    skipping fully-blank rows."""
    ws = wb[name]
    headers = [clean(c.value) for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in raw):
            continue
        row = {headers[i]: clean(raw[i]) for i in range(len(headers)) if headers[i]}
        rows.append(row)
    return rows


def parse_int(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    m = re.match(r"^-?\d+", s.replace(",", ""))
    if not m:
        return None
    try:
        return int(m.group(0))
    except ValueError:
        return None


def parse_range_midpoint(v):
    """Parse '300-400' / '1,000-1,200' / '1,632' to an int midpoint.
    Rejects anything containing letters (percentiles, footnotes, etc.)."""
    if v is None:
        return None
    s = str(v).strip()
    if re.search(r"[A-Za-z]", s):
        return None
    nums = re.findall(r"[\d,]+", s)
    if not nums:
        return None
    vals = [int(n.replace(",", "")) for n in nums]
    if len(vals) == 1:
        return vals[0]
    if len(vals) == 2:
        return round((vals[0] + vals[1]) / 2)
    return None


# ───────────────────────── colleges ─────────────────────────


def make_short_name(name):
    s = name
    subs = [
        (r"\bGovernment Women\b", "Govt. Women"),
        (r"\bGovernment\b", "Govt."),
        (r"\bPolytechnic College\b", "Poly"),
        (r"\bEngineering College\b", "Engg. College"),
        (r"\bInstitute of Engineering (&|and) Technology\b", "IET"),
        (r"\bInstitute of Technology\b", "IT"),
        (r"\bCollege of Engineering\b", "CoE"),
    ]
    for pat, rep in subs:
        s = re.sub(pat, rep, s)
    s = s.replace(",", "")
    s = _WHITESPACE.sub(" ", s).strip()
    return s


def fee_label(fee):
    if fee is None:
        return None
    if fee >= 100000:
        return "₹{:.1f}L".format(fee / 100000)
    return "₹{}k".format(round(fee / 1000))


def four_year_total_label(fee):
    if fee is None:
        return None
    return "₹{:.1f} L".format(fee * 4 / 100000)


def build_colleges(wb):
    inst_rows = sheet_rows(wb, "02_Institutions_Master")
    poly_rows = sheet_rows(wb, "03_Polytechnic_Current")

    poly_by_id = {}
    for r in poly_rows:
        iid = r.get("Institution_ID")
        if iid and iid not in poly_by_id:
            poly_by_id[iid] = r

    colleges = []
    for r in inst_rows:
        iid = r.get("Institution_ID")
        if not iid:
            continue
        level_raw = r.get("Institution_Level") or ""
        level = "Diploma" if "Diploma" in level_raw else "Engineering"
        type_raw = r.get("Institution_Type") or ""
        ctype = INSTITUTION_TYPE_MAP.get(type_raw, type_raw)
        poly = poly_by_id.get(iid)
        established = parse_int(poly.get("Establish_Year")) if poly else None

        colleges.append(
            {
                "id": iid,
                "name": r.get("Institution_Name"),
                "shortName": make_short_name(r.get("Institution_Name") or ""),
                "level": level,
                "type": ctype,
                "city": r.get("Location"),
                "district": r.get("District"),
                "affiliation": r.get("Affiliation_or_University"),
                "established": established,
                "fee": None,
                "feeLabel": None,
                "fourYearTotal": None,
                "placed": None,
                "seats": None,
                "sourceUrl": r.get("Official_Source_URL"),
                "sourceStatus": norm_status(r.get("Source_Status")),
            }
        )

    colleges.sort(key=lambda c: c["id"])
    return colleges


def match_college(inst_name, colleges_by_norm_name, colleges_by_norm_name_city, colleges):
    n = norm(inst_name)
    if n in colleges_by_norm_name:
        return colleges_by_norm_name[n]
    if n in colleges_by_norm_name_city:
        return colleges_by_norm_name_city[n]
    return None


def attach_fees(colleges, wb):
    rows = sheet_rows(wb, "08_Fees")
    rows.sort(key=lambda r: r.get("Record_ID") or "")

    by_norm_name = {}
    by_norm_name_city = {}
    for c in colleges:
        by_norm_name.setdefault(norm(c["name"]), c)
        by_norm_name_city.setdefault(norm("{} {}".format(c["name"], c["city"] or "")), c)

    assigned = set()
    for r in rows:
        college = match_college(r.get("Institution"), by_norm_name, by_norm_name_city, colleges)
        if college is None or college["id"] in assigned:
            continue
        if (r.get("Currency") or "").upper() != "INR":
            continue
        amount = parse_int(r.get("Amount"))
        if amount is None or amount <= 0:
            continue
        college["fee"] = amount
        college["feeLabel"] = fee_label(amount)
        college["fourYearTotal"] = four_year_total_label(amount)
        assigned.add(college["id"])


def attach_placements(colleges, wb):
    rows = sheet_rows(wb, "11_Placements")

    by_norm_name = {}
    by_norm_name_city = {}
    for c in colleges:
        by_norm_name.setdefault(norm(c["name"]), c)
        by_norm_name_city.setdefault(norm("{} {}".format(c["name"], c["city"] or "")), c)

    # group by matched college id, keep the most recent Year_or_Period
    best = {}
    for r in rows:
        college = match_college(r.get("Institution"), by_norm_name, by_norm_name_city, colleges)
        if college is None:
            continue
        placed = parse_int(r.get("Placed_Count"))
        if placed is None:
            continue
        higher = parse_int(r.get("Higher_Study_Count")) or 0
        self_emp = parse_int(r.get("Self_Employed_Count")) or 0
        period = r.get("Year_or_Period") or ""
        year_match = re.match(r"(\d{4})", period)
        year_key = int(year_match.group(1)) if year_match else -1
        denom = placed + higher + self_emp
        if denom <= 0:
            continue
        pct = round(placed / denom * 100)
        entry = (year_key, r.get("Record_ID") or "")
        prev = best.get(college["id"])
        if prev is None or entry > prev[0]:
            best[college["id"]] = (entry, pct)

    for c in colleges:
        if c["id"] in best:
            c["placed"] = best[c["id"]][1]


def attach_seats(colleges, wb):
    rows = sheet_rows(wb, "05_Programs_Courses")

    by_norm_name = {}
    by_norm_name_city = {}
    for c in colleges:
        by_norm_name.setdefault(norm(c["name"]), c)
        by_norm_name_city.setdefault(norm("{} {}".format(c["name"], c["city"] or "")), c)

    totals = {}
    for r in rows:
        college = match_college(r.get("Institution"), by_norm_name, by_norm_name_city, colleges)
        if college is None:
            continue
        intake = parse_int(r.get("Intake"))
        if intake is None:
            continue
        totals[college["id"]] = totals.get(college["id"], 0) + intake

    for c in colleges:
        if c["id"] in totals:
            c["seats"] = totals[c["id"]]


# ───────────────────────── cutoffs ─────────────────────────


def collect_real_anchors(wb):
    """Real (collegeId, branch) -> (closing:int, source:'PRIMARY'|'SECONDARY')
    from 07_Cutoffs, General category only. Deduplicated by ascending
    Record_ID (first row wins) for determinism."""
    rows = sheet_rows(wb, "07_Cutoffs")
    rows.sort(key=lambda r: r.get("Record_ID") or "")

    anchors = {}
    for r in rows:
        if (r.get("Category") or "") != "General":
            continue
        college_label = r.get("College") or ""
        college_id = CUTOFF_COLLEGE_ALIAS.get(college_label)
        if college_id is None:
            continue
        branch = CUTOFF_BRANCH_ALIAS.get(r.get("Branch") or "")
        if branch is None:
            continue
        midpoint = parse_range_midpoint(r.get("Cutoff_or_Range"))
        if midpoint is None or midpoint <= 0:
            continue
        key = (college_id, branch)
        if key in anchors:
            continue  # first (lowest Record_ID) wins
        source = norm_status(r.get("Authority"))
        anchors[key] = (midpoint, source)
    return anchors


def build_cutoffs(colleges, real_anchors):
    cutoffs = []
    for college in colleges:
        college_id = college["id"]
        ctype = college["type"]
        base = TYPE_BASE.get(ctype, TYPE_BASE["Government"])
        crc = zlib.crc32(college_id.encode("utf-8"))
        offset_frac = (crc % 1000) / 999.0 * 0.70 - 0.35
        type_anchor = base * (1 + offset_frac)

        cse_real = real_anchors.get((college_id, "CSE"))
        if cse_real is not None:
            cse_anchor, cse_source = cse_real
        else:
            cse_anchor, cse_source = type_anchor, "DEMO"

        for branch in BRANCHES:
            if branch == "CSE":
                branch_general = cse_anchor
                branch_general_source = cse_source
            else:
                real = real_anchors.get((college_id, branch))
                if real is not None and real[0] > cse_anchor:
                    branch_general, branch_general_source = real
                else:
                    branch_general = cse_anchor * BRANCH_FACTOR[branch]
                    branch_general_source = "DEMO"

            for category in CATEGORIES:
                mult = CATEGORY_MULTIPLIER[category]
                closing = int(round(branch_general * mult / 10.0)) * 10
                if closing <= 0:
                    closing = 10
                source = branch_general_source if category == "General" else "DEMO"
                cutoffs.append(
                    {
                        "collegeId": college_id,
                        "branch": branch,
                        "category": category,
                        "year": 2025,
                        "closing": closing,
                        "source": source,
                    }
                )

    cutoffs.sort(key=lambda c: (c["collegeId"], BRANCHES.index(c["branch"]), CATEGORIES.index(c["category"])))
    return cutoffs


# ───────────────────────── straight pass-through sheets ─────────────────────────


def build_fees(wb):
    rows = sheet_rows(wb, "08_Fees")
    out = []
    for r in rows:
        out.append(
            {
                "id": r.get("Record_ID"),
                "institution": r.get("Institution"),
                "program": r.get("Program/Context"),
                "academicYear": r.get("Academic_Year"),
                "component": r.get("Fee_Component"),
                "amount": parse_int(r.get("Amount")),
                "currency": r.get("Currency"),
                "eligibility": r.get("Eligibility/Category"),
                "sourceStatus": norm_status(r.get("Source_Status")),
                "sourceUrl": r.get("Source_URL"),
                "notes": r.get("Notes"),
            }
        )
    out.sort(key=lambda x: x["id"] or "")
    return out


def build_scholarships(wb):
    rows = sheet_rows(wb, "09_Scholarships")
    out = []
    for r in rows:
        out.append(
            {
                "id": r.get("Record_ID"),
                "scheme": r.get("Scheme_or_Resource"),
                "target": r.get("Target/Context"),
                "academicYear": r.get("Academic_Year"),
                "info": r.get("Key_Information"),
                "status": r.get("Status"),
                "sourceUrl": r.get("Source_URL"),
                "chatbotUse": r.get("Chatbot_Use"),
                "caveat": r.get("Caveat"),
            }
        )
    out.sort(key=lambda x: x["id"] or "")
    return out


def build_placements(wb):
    rows = sheet_rows(wb, "11_Placements")
    out = []
    for r in rows:
        out.append(
            {
                "id": r.get("Record_ID"),
                "institution": r.get("Institution"),
                "period": r.get("Year_or_Period"),
                "branch": r.get("Branch_or_Context"),
                "placedCount": parse_int(r.get("Placed_Count")),
                "higherStudyCount": parse_int(r.get("Higher_Study_Count")),
                "selfEmployedCount": parse_int(r.get("Self_Employed_Count")),
                "package": r.get("Salary_or_Package"),
                "majorEmployers": r.get("Major_Employers"),
                "sourceStatus": norm_status(r.get("Source_Status")),
                "sourceUrl": r.get("Source_URL"),
                "notes": r.get("Notes"),
            }
        )
    out.sort(key=lambda x: x["id"] or "")
    return out


def build_programs(wb):
    rows = sheet_rows(wb, "05_Programs_Courses")
    out = []
    for r in rows:
        out.append(
            {
                "id": r.get("Record_ID"),
                "institution": r.get("Institution"),
                "level": r.get("Level"),
                "branch": r.get("Program/Branch"),
                "intake": parse_int(r.get("Intake")),
                "academicYear": r.get("Academic_Year"),
                "sourceStatus": norm_status(r.get("Source_Status")),
                "sourceUrl": r.get("Source_URL"),
                "notes": r.get("Notes"),
            }
        )
    out.sort(key=lambda x: x["id"] or "")
    return out


def build_faqs(wb):
    rows = sheet_rows(wb, "16_Chatbot_FAQ_Intents")
    out = []
    for r in rows:
        out.append(
            {
                "id": r.get("FAQ_ID"),
                "question": r.get("User_Intent_or_Question"),
                "shortAnswer": r.get("Short_Answer"),
                "detail": r.get("Detail_or_Steps"),
                "sourceSheet": r.get("Answer_Source_Sheet"),
                "sourceStatus": r.get("Source_Status"),
                "caveat": r.get("Caveat_or_Disclaimer"),
            }
        )
    out.sort(key=lambda x: x["id"] or "")
    return out


def build_glossary(wb):
    rows = sheet_rows(wb, "17_Glossary")
    out = []
    for r in rows:
        out.append(
            {
                "term": r.get("Term"),
                "fullForm": r.get("Full_Form"),
                "meaning": r.get("Plain-Language_Meaning"),
                "relevance": r.get("Relevance_to_User"),
            }
        )
    out.sort(key=lambda x: x["term"] or "")
    return out


def build_contacts(wb):
    rows = sheet_rows(wb, "13_Contacts_Notices")
    out = []
    for r in rows:
        out.append(
            {
                "id": r.get("Record_ID"),
                "type": r.get("Type"),
                "office": r.get("Office_or_Institution"),
                "purpose": r.get("Purpose"),
                "contact": r.get("Contact_or_Notice"),
                "date": r.get("Date"),
                "status": r.get("Status"),
                "sourceUrl": r.get("Source_URL"),
                "action": r.get("Chatbot_Action"),
            }
        )
    out.sort(key=lambda x: x["id"] or "")
    return out


# ───────────────────────── JS emission ─────────────────────────


def to_js_literal(value, indent=0):
    pad = "  " * indent
    pad_in = "  " * (indent + 1)
    if isinstance(value, dict):
        if not value:
            return "{}"
        items = []
        for k in value.keys():
            v = to_js_literal(value[k], indent + 1)
            items.append("{}{}: {}".format(pad_in, json.dumps(k, ensure_ascii=False), v))
        return "{\n" + ",\n".join(items) + "\n" + pad + "}"
    if isinstance(value, list):
        if not value:
            return "[]"
        items = [pad_in + to_js_literal(v, indent + 1) for v in value]
        return "[\n" + ",\n".join(items) + "\n" + pad + "]"
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return "null"
    if isinstance(value, (int, float)):
        return json.dumps(value)
    return json.dumps(value, ensure_ascii=False)


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

    colleges = build_colleges(wb)
    attach_fees(colleges, wb)
    attach_placements(colleges, wb)
    attach_seats(colleges, wb)

    real_anchors = collect_real_anchors(wb)
    cutoffs = build_cutoffs(colleges, real_anchors)

    fees = build_fees(wb)
    scholarships = build_scholarships(wb)
    placements = build_placements(wb)
    programs = build_programs(wb)
    faqs = build_faqs(wb)
    glossary = build_glossary(wb)
    contacts = build_contacts(wb)

    diploma_count = sum(1 for c in colleges if c["level"] == "Diploma")
    engineering_count = sum(1 for c in colleges if c["level"] == "Engineering")

    data = {
        "colleges": colleges,
        "cutoffs": cutoffs,
        "fees": fees,
        "scholarships": scholarships,
        "placements": placements,
        "programs": programs,
        "faqs": faqs,
        "glossary": glossary,
        "contacts": contacts,
        "meta": {
            "collegeCount": len(colleges),
            "diplomaCount": diploma_count,
            "engineeringCount": engineering_count,
            "verifiedOn": VERIFIED_ON,
        },
    }

    body = to_js_literal(data, indent=0)

    js = (
        "// GENERATED FILE — do not hand-edit.\n"
        "// Produced by tools/build-data.py from 2026.08-College-Database-final.xlsx\n"
        "var __g = typeof window !== 'undefined' ? window : globalThis;\n"
        "__g.UNIVERSE_DATA = " + body + ";\n"
        "\n"
        "if (typeof module !== 'undefined' && module.exports) module.exports = __g.UNIVERSE_DATA;\n"
    )

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8", newline="\n") as f:
        f.write(js)

    print("Wrote {} ({} colleges, {} cutoffs)".format(OUT_PATH, len(colleges), len(cutoffs)))


if __name__ == "__main__":
    main()
